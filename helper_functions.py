"""
Vehicle Repair Estimate Auto-Approval Analysis
================================================
Enhanced version: correlation-first variable selection, exhaustive multi-variable
rule search (up to 5 variables), and full Excel output.

Usage
-----
1. Update DATA_PATH to your CSV / DB export.
2. Run:  python auto_approval_analysis.py
3. Open: auto_approval_analysis.xlsx
"""

import pandas as pd
import numpy as np
import warnings
from itertools import combinations
from pathlib import Path

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# 0. CONFIGURATION  ← update these paths
# ─────────────────────────────────────────────
DATA_PATH   = "your_estimates_file.csv"   # <-- change to your file
OUTPUT_PATH = "auto_approval_analysis.xlsx"

MIN_VENDOR_ESTIMATES   = 10   # minimum estimates before a vendor gets a tier
MIN_RULE_COVERAGE      = 0.02 # rules must cover at least 2% of estimates
MAX_RULE_ERROR_RATE    = 0.25 # rules must have ≤ 25% wrong approvals
TOP_N_RULES            = 50   # keep the top-N rules by score in Excel
MAX_COMBO_VARS         = 5    # maximum variables to combine in a rule (1-5)

# ─────────────────────────────────────────────
# 1. LOAD DATA
# ─────────────────────────────────────────────
print("── 1. Loading data ──────────────────────")
df = pd.read_csv(DATA_PATH, low_memory=False)
print(f"  Loaded {len(df):,} rows, {df.shape[1]} columns")

# ─────────────────────────────────────────────
# 2. DATA CLEANING
# ─────────────────────────────────────────────
print("\n── 2. Cleaning ──────────────────────────")

# Drop columns with no signal (single unique value or near-total nulls)
DROP_COLS = [
    "is_glass_est_ind",       # all zeros
    "temp_est_ind",           # all zeros
    "is_bulk_ind",            # single value
    "is_electronic_est_ind",  # single value
    "cdr_vndr_flag",          # single value
    "expd_cmpl_dte",          # 99.99% null
    "slvg_amt",               # all null
    "email_txt",              # all null
]
df.drop(columns=[c for c in DROP_COLS if c in df.columns], inplace=True)

if "time_to_approve_days" in df.columns:
    df["time_to_approve_days"] = df["time_to_approve_days"].clip(lower=0)

for col in ["est_recv_dte", "apprv_dte", "act_cmpl_dte"]:
    if col in df.columns:
        df[col] = pd.to_datetime(df[col], errors="coerce")

# Also drop low-variance columns automatically
for col in df.select_dtypes(include=[np.number]).columns:
    if df[col].nunique() <= 1:
        df.drop(columns=[col], inplace=True)

print(f"  Cleaned shape: {df.shape}")

# ─────────────────────────────────────────────
# 3. TARGET VARIABLE
# ─────────────────────────────────────────────
print("\n── 3. Target variable ──────────────────")
df["first_pass"] = (df["rvsn_nbr"] == 1).astype(int)
overall_rate = df["first_pass"].mean()
print(f"  Overall first-pass rate : {overall_rate:.1%}")
print(f"  First-pass approvals    : {df['first_pass'].sum():,}")
print(f"  Needed revisions        : {(df['first_pass']==0).sum():,}")

# ─────────────────────────────────────────────
# 4. FEATURE ENGINEERING
# ─────────────────────────────────────────────
print("\n── 4. Feature engineering ──────────────")

if "lbr_hr_qty" in df.columns and "est_tot_amt" in df.columns:
    df["cost_per_lbr_hr"] = np.where(
        df["lbr_hr_qty"] > 0,
        df["est_tot_amt"] / df["lbr_hr_qty"],
        np.nan
    )

if "veh_yr" in df.columns:
    df["veh_age"] = pd.Timestamp.now().year - df["veh_yr"]

# Vendor historical approval rate (leave-one-out risk is small at this scale)
if "vr_vndr_id" in df.columns:
    vendor_rates = (
        df.groupby("vr_vndr_id")["first_pass"]
        .agg(vendor_approval_rate="mean", vendor_est_count="count")
        .reset_index()
    )
    vendor_rates["vendor_tier"] = np.where(
        vendor_rates["vendor_est_count"] < MIN_VENDOR_ESTIMATES,
        "insufficient_history",
        pd.qcut(
            vendor_rates["vendor_approval_rate"],
            q=[0, 0.25, 0.5, 0.75, 1.0],
            labels=["low", "below_avg", "above_avg", "trusted"],
            duplicates="drop"
        ).astype(str)
    )
    df = df.merge(
        vendor_rates[["vr_vndr_id", "vendor_approval_rate", "vendor_tier"]],
        on="vr_vndr_id", how="left"
    )

# Parts-to-labour ratio
if "est_tot_amt" in df.columns and "lbr_hr_qty" in df.columns:
    lbr_cost_est = df["lbr_hr_qty"] * df.get("bdy_lbr_rate", pd.Series(50, index=df.index))
    df["parts_to_labour_ratio"] = np.where(
        lbr_cost_est > 0,
        (df["est_tot_amt"] - lbr_cost_est) / lbr_cost_est,
        np.nan
    )

print("  Engineered: cost_per_lbr_hr, veh_age, vendor_approval_rate,")
print("              vendor_tier, parts_to_labour_ratio")

# ─────────────────────────────────────────────
# 5. CORRELATION ANALYSIS — pick top variables
# ─────────────────────────────────────────────
print("\n── 5. Correlation analysis ─────────────")

CANDIDATE_NUMERIC = [
    "est_tot_amt", "lbr_hr_qty", "line_item_count", "cost_per_lbr_hr",
    "time_to_approve_days", "vendor_approval_rate", "veh_age",
    "bdy_lbr_rate", "mchncl_lbr_rate", "frm_lbr_rate", "pnt_mtrl_rate",
    "dmstc_part_disc_amt", "frn_part_disc_amt", "parts_to_labour_ratio",
]
numeric_cols = [c for c in CANDIDATE_NUMERIC if c in df.columns]

corr_df = (
    df[numeric_cols + ["first_pass"]]
    .corr()["first_pass"]
    .drop("first_pass")
    .sort_values(key=abs, ascending=False)
    .reset_index()
    .rename(columns={"index": "feature", "first_pass": "correlation"})
)
corr_df["abs_corr"] = corr_df["correlation"].abs()
corr_df["correlation"] = corr_df["correlation"].round(4)

print("  Top numeric correlations with first_pass:")
print(corr_df.head(10).to_string(index=False))

# Select top-N numeric features by absolute correlation (exclude vendor_approval_rate
# since vendor_tier encodes it categorically; keep it too for reference)
TOP_NUMERIC = corr_df.head(8)["feature"].tolist()

# ─────────────────────────────────────────────
# 6. UNIVARIATE BUCKET ANALYSIS
# ─────────────────────────────────────────────
print("\n── 6. Univariate bucket analysis ───────")

def approval_summary(data, group_col):
    g = (
        data.groupby(group_col)["first_pass"]
        .agg(approval_rate="mean", total_estimates="count", first_pass_count="sum")
        .reset_index()
    )
    g["revision_count"]    = g["total_estimates"] - g["first_pass_count"]
    g["approval_rate_pct"] = (g["approval_rate"] * 100).round(1)
    g["coverage_pct"]      = (g["total_estimates"] / len(df) * 100).round(1)
    return g.sort_values("approval_rate", ascending=False)

# Amount buckets
bucket_uni_results = {}
if "est_tot_amt" in df.columns:
    df["est_amt_bucket"] = pd.cut(
        df["est_tot_amt"],
        bins=[0,250,500,750,1000,1500,2500,np.inf],
        labels=["$0–250","$251–500","$501–750","$751–1k","$1k–1.5k","$1.5k–2.5k","$2.5k+"],
        right=True
    )
    bucket_uni_results["est_amt_bucket"] = approval_summary(df, "est_amt_bucket")

if "lbr_hr_qty" in df.columns:
    df["lbr_hr_bucket"] = pd.cut(
        df["lbr_hr_qty"],
        bins=[0,2,4,8,16,np.inf],
        labels=["0–2 hrs","2–4 hrs","4–8 hrs","8–16 hrs","16+ hrs"],
        right=True
    )
    bucket_uni_results["lbr_hr_bucket"] = approval_summary(df, "lbr_hr_bucket")

if "vendor_tier" in df.columns:
    bucket_uni_results["vendor_tier"] = approval_summary(df, "vendor_tier")

if "licplte_st" in df.columns:
    bucket_uni_results["licplte_st"] = approval_summary(df, "licplte_st")

# ─────────────────────────────────────────────
# 7. THRESHOLD GRID — per top numeric variable
# ─────────────────────────────────────────────
print("\n── 7. Per-variable threshold grids ─────")

def make_threshold_grid(col, n_thresholds=20):
    """For a numeric column, sweep thresholds and record approval rate + coverage."""
    series = df[col].dropna()
    thresholds = np.percentile(series, np.linspace(5, 95, n_thresholds))
    rows = []
    for t in thresholds:
        mask   = df[col] <= t
        subset = df[mask]
        if len(subset) == 0:
            continue
        rows.append({
            "variable"         : col,
            "threshold"        : round(t, 2),
            "coverage_pct"     : round(len(subset)/len(df)*100, 1),
            "approval_rate_pct": round(subset["first_pass"].mean()*100, 1),
            "first_pass_count" : int(subset["first_pass"].sum()),
            "total_count"      : len(subset),
            "error_rate_pct"   : round((subset["first_pass"]==0).mean()*100, 1),
        })
    return pd.DataFrame(rows)

threshold_grids = []
for col in TOP_NUMERIC:
    if col in df.columns and df[col].notna().sum() > 100:
        threshold_grids.append(make_threshold_grid(col))

threshold_grid_df = pd.concat(threshold_grids, ignore_index=True) if threshold_grids else pd.DataFrame()

# ─────────────────────────────────────────────
# 8. EXHAUSTIVE MULTI-VARIABLE RULE SEARCH
#    (combinations of 1–5 variables)
# ─────────────────────────────────────────────
print("\n── 8. Multi-variable rule search ───────")
print(f"  Variables in search pool: {TOP_NUMERIC}")

# Build percentile thresholds for each variable (fewer points = manageable combos)
N_THRESH_PER_VAR = 5   # 5th, 25th, 50th, 75th, 95th percentiles

def get_thresholds(col):
    s = df[col].dropna()
    return np.unique(np.percentile(s, [5, 25, 50, 75, 95])).tolist()

var_thresholds = {}
for col in TOP_NUMERIC:
    if col in df.columns and df[col].notna().sum() > 100:
        var_thresholds[col] = get_thresholds(col)

search_vars = list(var_thresholds.keys())

rule_records = []
n_evaluated  = 0

for n_vars in range(1, MAX_COMBO_VARS + 1):
    for var_combo in combinations(search_vars, n_vars):
        # build all threshold combinations for this variable set
        thresh_options = [var_thresholds[v] for v in var_combo]
        from itertools import product as iproduct
        for thresh_combo in iproduct(*thresh_options):
            n_evaluated += 1

            # build mask: all selected variables ≤ their threshold
            mask = pd.Series(True, index=df.index)
            for var, thr in zip(var_combo, thresh_combo):
                mask &= df[var].fillna(np.inf) <= thr

            subset = df[mask]
            coverage = len(subset) / len(df)
            if len(subset) == 0 or coverage < MIN_RULE_COVERAGE:
                continue

            error_rate = (subset["first_pass"] == 0).mean()
            if error_rate > MAX_RULE_ERROR_RATE:
                continue

            approval_rate = subset["first_pass"].mean()

            # Score: reward high approval rate and wide coverage
            score = approval_rate * np.log1p(coverage * 100)

            conditions = " AND ".join(
                [f"{v} ≤ {t:,.2f}" for v, t in zip(var_combo, thresh_combo)]
            )
            rule_records.append({
                "n_vars"           : n_vars,
                "rule_conditions"  : conditions,
                "coverage_pct"     : round(coverage * 100, 1),
                "approval_rate_pct": round(approval_rate * 100, 1),
                "error_rate_pct"   : round(error_rate * 100, 1),
                "total_covered"    : len(subset),
                "correct_approvals": int(subset["first_pass"].sum()),
                "wrong_approvals"  : int((subset["first_pass"]==0).sum()),
                "score"            : round(score, 4),
            })

print(f"  Rules evaluated    : {n_evaluated:,}")
print(f"  Rules passing filters: {len(rule_records):,}")

rules_df = (
    pd.DataFrame(rule_records)
    .sort_values("score", ascending=False)
    .drop_duplicates(subset="rule_conditions")
    .head(TOP_N_RULES)
    .reset_index(drop=True)
)
rules_df.index += 1
rules_df.index.name = "rank"
rules_df = rules_df.reset_index()

print(f"\n  Top 10 rules by score:")
print(rules_df.head(10)[
    ["rank","n_vars","rule_conditions","coverage_pct","approval_rate_pct","error_rate_pct"]
].to_string(index=False))

# ─────────────────────────────────────────────
# 9. BEST RULE PER n_vars
# ─────────────────────────────────────────────
best_per_n = (
    rules_df.sort_values("score", ascending=False)
    .groupby("n_vars", as_index=False)
    .first()
    .sort_values("n_vars")
)
print("\n  Best rule per number of variables:")
print(best_per_n[
    ["n_vars","rule_conditions","coverage_pct","approval_rate_pct","error_rate_pct","score"]
].to_string(index=False))

# ─────────────────────────────────────────────
# 10. VENDOR DEEP-DIVE
# ─────────────────────────────────────────────
if "vr_vndr_id" in df.columns:
    vendor_detail = (
        df.groupby("vr_vndr_id").agg(
            total_estimates   = ("first_pass", "count"),
            first_pass_count  = ("first_pass", "sum"),
            approval_rate     = ("first_pass", "mean"),
            avg_est_amt       = ("est_tot_amt", "mean"),
            avg_lbr_hrs       = ("lbr_hr_qty",  "mean"),
            avg_line_items    = ("line_item_count", "mean"),
            avg_time_to_approve = ("time_to_approve_days", "mean"),
        )
        .query("total_estimates >= @MIN_VENDOR_ESTIMATES")
        .assign(
            approval_rate_pct = lambda x: (x["approval_rate"] * 100).round(1),
            avg_est_amt       = lambda x: x["avg_est_amt"].round(0),
            avg_lbr_hrs       = lambda x: x["avg_lbr_hrs"].round(1),
            avg_line_items    = lambda x: x["avg_line_items"].round(1),
        )
        .sort_values("approval_rate", ascending=False)
        .reset_index()
    )
else:
    vendor_detail = pd.DataFrame()

# ─────────────────────────────────────────────
# 11. CROSS-TAB HEATMAP
# ─────────────────────────────────────────────
cross_tab_rate  = pd.DataFrame()
cross_tab_count = pd.DataFrame()
if "est_amt_bucket" in df.columns and "lbr_hr_bucket" in df.columns:
    cross_tab = (
        df.groupby(["est_amt_bucket", "lbr_hr_bucket"])["first_pass"]
        .agg(approval_rate="mean", count="count")
        .unstack("lbr_hr_bucket")
        .round(3)
    )
    cross_tab_rate  = (cross_tab["approval_rate"] * 100).round(1)
    cross_tab_count = cross_tab["count"]

# ─────────────────────────────────────────────
# 12. OPTIONAL DECISION TREE
# ─────────────────────────────────────────────
print("\n── 12. Decision tree rule extraction ───")
try:
    from sklearn.tree import DecisionTreeClassifier, export_text

    feature_cols = [c for c in TOP_NUMERIC if c in df.columns]
    X = df[feature_cols].copy()
    y = df["first_pass"]
    mask = X.notna().all(axis=1) & y.notna()
    X, y = X[mask], y[mask]

    clf = DecisionTreeClassifier(
        max_depth=5,
        min_samples_leaf=50,
        class_weight="balanced",
        random_state=42
    )
    clf.fit(X, y)
    tree_rules = export_text(clf, feature_names=feature_cols)
    print(tree_rules[:2000])

    feature_importance = pd.DataFrame({
        "feature"   : feature_cols,
        "importance": clf.feature_importances_
    }).sort_values("importance", ascending=False)
    dt_available = True
except ImportError:
    print("  scikit-learn not installed. Skipping.")
    tree_rules         = "scikit-learn not available"
    feature_importance = pd.DataFrame()
    dt_available       = False

# ─────────────────────────────────────────────
# 13. EXPORT TO EXCEL
# ─────────────────────────────────────────────
print(f"\n── 13. Writing Excel → {OUTPUT_PATH}")

from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

HEADER_FILL  = PatternFill("solid", start_color="1F4E79")  # dark blue
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT    = Font(name="Arial", size=10)
ALT_FILL     = PatternFill("solid", start_color="D6E4F0")   # light blue alternate row
BEST_FILL    = PatternFill("solid", start_color="E2EFDA")   # light green highlight
TITLE_FONT   = Font(bold=True, name="Arial", size=12, color="1F4E79")
thin         = Side(style="thin", color="AAAAAA")
BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_sheet(ws, df_rows, freeze="A2", col_widths=None):
    """Apply header style + alternating rows to a worksheet."""
    # Header row
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
    # Data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.font      = BODY_FONT
            cell.alignment = Alignment(vertical="center")
            cell.fill      = fill
            cell.border    = BORDER
    # Column widths
    if col_widths:
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
    else:
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    ws.row_dimensions[1].height = 30
    if freeze:
        ws.freeze_panes = freeze

with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:

    # ── Sheet 1: Summary ────────────────────────────────────────────────
    summary_data = pd.DataFrame({
        "Metric": [
            "Total estimates",
            "First-pass approvals",
            "Needed revisions",
            "Overall first-pass rate",
            "Unique vendors",
            "Estimate amount range",
            "Labour hours range",
            "Top correlated variable",
            "Top correlated variable (r)",
        ],
        "Value": [
            f"{len(df):,}",
            f"{df['first_pass'].sum():,}",
            f"{(df['first_pass']==0).sum():,}",
            f"{overall_rate:.1%}",
            f"{df['vr_vndr_id'].nunique():,}" if 'vr_vndr_id' in df.columns else "N/A",
            f"${df['est_tot_amt'].min():.0f} – ${df['est_tot_amt'].max():,.0f}" if 'est_tot_amt' in df.columns else "N/A",
            f"{df['lbr_hr_qty'].min():.0f} – {df['lbr_hr_qty'].max():.0f} hrs" if 'lbr_hr_qty' in df.columns else "N/A",
            corr_df.iloc[0]["feature"]      if len(corr_df) else "N/A",
            f"{corr_df.iloc[0]['correlation']:.4f}" if len(corr_df) else "N/A",
        ]
    })
    summary_data.to_excel(writer, sheet_name="Summary", index=False)

    # ── Sheet 2: Correlations ────────────────────────────────────────────
    corr_df.to_excel(writer, sheet_name="Correlations", index=False)

    # ── Sheet 3: Threshold Grids (per variable) ──────────────────────────
    if not threshold_grid_df.empty:
        threshold_grid_df.to_excel(writer, sheet_name="Threshold_Grids", index=False)

    # ── Sheet 4: Top Rules (all combos) ──────────────────────────────────
    rules_df.to_excel(writer, sheet_name="Top_Rules", index=False)

    # ── Sheet 5: Best Rule per N Variables ──────────────────────────────
    best_per_n.to_excel(writer, sheet_name="Best_Per_NVars", index=False)

    # ── Sheet 6: By Amount Bucket ────────────────────────────────────────
    if "est_amt_bucket" in bucket_uni_results:
        bucket_uni_results["est_amt_bucket"].to_excel(
            writer, sheet_name="By_Amount_Bucket", index=False)

    # ── Sheet 7: By Labour Hours ─────────────────────────────────────────
    if "lbr_hr_bucket" in bucket_uni_results:
        bucket_uni_results["lbr_hr_bucket"].to_excel(
            writer, sheet_name="By_Labour_Hours", index=False)

    # ── Sheet 8: By Vendor Tier ──────────────────────────────────────────
    if "vendor_tier" in bucket_uni_results:
        bucket_uni_results["vendor_tier"].to_excel(
            writer, sheet_name="By_Vendor_Tier", index=False)

    # ── Sheet 9: By State ────────────────────────────────────────────────
    if "licplte_st" in bucket_uni_results:
        bucket_uni_results["licplte_st"].to_excel(
            writer, sheet_name="By_State", index=False)

    # ── Sheet 10: Vendor Detail ──────────────────────────────────────────
    if not vendor_detail.empty:
        vendor_detail.to_excel(writer, sheet_name="Vendor_Detail", index=False)

    # ── Sheet 11: Cross-tab Approval Rate ───────────────────────────────
    if not cross_tab_rate.empty:
        cross_tab_rate.to_excel(writer, sheet_name="Crosstab_Approval_Rate")
    if not cross_tab_count.empty:
        cross_tab_count.to_excel(writer, sheet_name="Crosstab_Count")

    # ── Sheet 12: Decision Tree Feature Importance ──────────────────────
    if dt_available and not feature_importance.empty:
        feature_importance.to_excel(
            writer, sheet_name="DT_Feature_Importance", index=False)

    # ── Sheet 13: Full Cleaned Data ──────────────────────────────────────
    export_cols = [
        "est_id", "vr_vndr_id", "vndr_grp_nbr", "est_tot_amt", "lbr_hr_qty",
        "line_item_count", "rvsn_nbr", "first_pass",
        "cost_per_lbr_hr", "est_amt_bucket", "lbr_hr_bucket",
        "vendor_approval_rate", "vendor_tier",
        "time_to_approve_days", "licplte_st", "veh_yr", "veh_age",
        "veh_make", "veh_modl", "dmg_dsc",
    ]
    export_cols = [c for c in export_cols if c in df.columns]
    df[export_cols].to_excel(writer, sheet_name="Cleaned_Data", index=False)

# ── Post-process: styling ────────────────────────────────────────────────
wb = load_workbook(OUTPUT_PATH)

# Summary
ws = wb["Summary"]
style_sheet(ws, [], col_widths={"A": 35, "B": 30})
ws["A1"].value = "Auto-Approval Analysis — Summary"
ws.merge_cells("A1:B1")
ws["A1"].font      = TITLE_FONT
ws["A1"].alignment = Alignment(horizontal="center")

# Correlations — add conditional colour scale
if "Correlations" in wb.sheetnames:
    ws = wb["Correlations"]
    style_sheet(ws, [])
    last_row = ws.max_row
    ws.conditional_formatting.add(
        f"B2:B{last_row}",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="num",   mid_value=0, mid_color="FFFFFF",
            end_type="max",   end_color="63BE7B"
        )
    )

# Top Rules — highlight best rule per n_vars in green
if "Top_Rules" in wb.sheetnames:
    ws = wb["Top_Rules"]
    style_sheet(ws, [], col_widths={"A": 6, "B": 6, "C": 70, "D": 14,
                                     "E": 18, "F": 14, "G": 16, "H": 18,
                                     "I": 16, "J": 12})
    best_rule_conditions = set(best_per_n["rule_conditions"].tolist())
    for row in ws.iter_rows(min_row=2):
        cond_val = row[2].value  # "rule_conditions" is column C (index 2)
        if cond_val in best_rule_conditions:
            for cell in row:
                cell.fill = BEST_FILL

# Best Per NVars
if "Best_Per_NVars" in wb.sheetnames:
    ws = wb["Best_Per_NVars"]
    style_sheet(ws, [], col_widths={"A": 8, "B": 70, "C": 14, "D": 18,
                                     "E": 14, "F": 16, "G": 18, "H": 16, "I": 12})
    # Highlight highest-scoring row
    best_score_row = None
    best_score_val = -1
    for row in ws.iter_rows(min_row=2):
        try:
            score_val = float(row[-1].value)
            if score_val > best_score_val:
                best_score_val = score_val
                best_score_row = row
        except (TypeError, ValueError):
            pass
    if best_score_row:
        for cell in best_score_row:
            cell.fill = PatternFill("solid", start_color="FFD700")  # gold

# Style all remaining plain sheets
for sname in wb.sheetnames:
    if sname not in ("Summary", "Correlations", "Top_Rules",
                     "Best_Per_NVars", "Crosstab_Approval_Rate", "Crosstab_Count"):
        ws = wb[sname]
        style_sheet(ws, [])

# Crosstab — colour scale on approval rate
if "Crosstab_Approval_Rate" in wb.sheetnames:
    ws = wb["Crosstab_Approval_Rate"]
    style_sheet(ws, [], freeze="B2")
    last_row = ws.max_row
    last_col = get_column_letter(ws.max_column)
    ws.conditional_formatting.add(
        f"B2:{last_col}{last_row}",
        ColorScaleRule(
            start_type="min",  start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max",    end_color="63BE7B"
        )
    )

wb.save(OUTPUT_PATH)
print(f"\n  Saved: {OUTPUT_PATH}")

print("""
  Sheets written:
    1  Summary                — dataset overview + top correlated variable
    2  Correlations           — all numeric correlations with first_pass (colour-coded)
    3  Threshold_Grids        — per-variable sweep across percentile thresholds
    4  Top_Rules              — top {top_n} multi-variable rules by score (best-per-n in green)
    5  Best_Per_NVars         — single best rule for 1-, 2-, 3-, 4-, 5-variable combos (gold row = top)
    6  By_Amount_Bucket       — first-pass rate per estimate amount band
    7  By_Labour_Hours        — first-pass rate per labour hour band
    8  By_Vendor_Tier         — first-pass rate per vendor quality tier
    9  By_State               — first-pass rate by licence-plate state
   10  Vendor_Detail          — per-vendor stats (all vendors ≥{min_v} estimates)
   11  Crosstab_Approval_Rate — amount × labour hours heat-map (colour gradient)
   12  Crosstab_Count         — same heat-map, count view
   13  DT_Feature_Importance  — decision tree feature ranking (if sklearn available)
   14  Cleaned_Data           — full dataset with engineered features

── Analysis complete ────────────────────────────
""".format(top_n=TOP_N_RULES, min_v=MIN_VENDOR_ESTIMATES))
